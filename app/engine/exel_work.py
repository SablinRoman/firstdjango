import django
import openpyxl
import os
import sys

wb = openpyxl.load_workbook('./Baza.xlsx')
sheet = wb.get_sheet_by_name('Проживающие')

project_dir = "/home/firtsdjango/app/engine/engine/"  # Путь до файла settings
sys.path.append(project_dir)
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'engine.settings')  # Добавление переменной окуружения
django.setup()
from hostel.models import Student  # Примечание: импорт не работает, если находться на верху
from hostel.models import Room  # Примечание: импорт не работает, если находться на верху


def import_rooms():
    i = int(input("Start row ="))
    print()

    print('import room')
    while sheet.cell(row=i, column=1).value is not None:

        if not Room.objects.filter(room_numb=sheet.cell(row=i, column=1).value):
            hostel = Room()
            hostel.room_numb = sheet.cell(row=i, column=1).value
            hostel.save()
            print('Комната', hostel.room_numb, 'импортирована')
        i += 1


def import_students():
    i = int(input("Start row ="))
    print()

    status_list = ['мужское', 'женское', 'пусто', 'занято']
    while sheet.cell(row=i, column=1).value is not None:

        student = Student()

        student.room = Room.objects.get(room_numb=sheet.cell(row=i, column=1).value)

        if sheet.cell(row=i, column=2).value in status_list:
            student.name = ''
            student.bed_status = sheet.cell(row=i, column=2).value
        else:
            student.name = sheet.cell(row=i, column=2).value
            student.bed_status = sheet.cell(row=i, column=6).value

        student.faculty = sheet.cell(row=i, column=3).value
        student.form_studies = sheet.cell(row=i, column=4).value
        student.group = sheet.cell(row=i, column=5).value
        student.sex = sheet.cell(row=i, column=6).value
        student.mobile_number = sheet.cell(row=i, column=7).value
        student.notation = sheet.cell(row=i, column=8).value

        if sheet.cell(row=i, column=9).value != '':
            student.fluorography = True
        else:
            student.fluorography = False

        if sheet.cell(row=i, column=10).value != '':
            student.pediculosis = True
        else:
            student.pediculosis = False

        student.contract_number = sheet.cell(row=i, column=11).value
        student.agreement_date = sheet.cell(row=i, column=12).value
        student.registration = sheet.cell(row=i, column=13).value
        student.citizenship = sheet.cell(row=i, column=14).value
        student.date_of_birthday = sheet.cell(row=i, column=15).value
        student.place_of_birthday = sheet.cell(row=i, column=16).value
        student.document_number = sheet.cell(row=i, column=17).value
        student.authority = sheet.cell(row=i, column=18).value
        student.date_of_issue = sheet.cell(row=i, column=19).value
        print(i, ' ', 'room=', student.room, ' ', student.name)
        student.save()

        i += 1
    print('Import completed')
